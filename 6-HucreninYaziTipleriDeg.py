import openpyxl
from openpyxl.styles import Font

calisamaKitabi = openpyxl.Workbook()
sayfaYapım = calisamaKitabi.active

sayfaYapım.cell(row=1,column=1).value = "Python Dili"
sayfaYapım.cell(row=1,column=1).font = Font(size=25)

sayfaYapım.cell(row=2,column=2).value="Python dili"
sayfaYapım.cell(row=2, column=2).font=Font(size=25, italic=True)

sayfaYapım.cell(row=3,column=3).value="Python dili"
sayfaYapım.cell(row=3,column=3).font=Font(size=25, bold=True)

sayfaYapım.cell(row=4,column=4).value="Python dili"
sayfaYapım.cell(row=4, column=4).font=Font(size=25, name='Times New Roman')

calisamaKitabi.save('yaziTipi.xlsx')