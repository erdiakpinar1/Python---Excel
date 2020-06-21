import openpyxl

calisamaKitabi = openpyxl.Workbook()
sayfaYapım = calisamaKitabi.active

sayfaYapım.cell(row=3, column=3).value = "Python"
sayfaYapım.cell(row=10, column=10).value = "excel"

sayfaYapım.row_dimensions[3].height=55
sayfaYapım.column_dimensions["J"].width=25
calisamaKitabi.save("boyutlandirma1.xlsx")