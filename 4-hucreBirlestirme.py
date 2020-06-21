import openpyxl

calisamaKitabi = openpyxl.Workbook()
sayfaYapım = calisamaKitabi.active

sayfaYapım.merge_cells('A2:C5')
sayfaYapım.merge_cells('D5:F6')

calisamaKitabi.save("Hucrebirlestirme.xlsx")